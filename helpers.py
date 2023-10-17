import os
import pathlib

import numpy as np
import openpyxl
import pandas as pd

def read_config_map(path):

    map_config  = {}
    dataframe = openpyxl.load_workbook(path)
 
    # Define variable to read sheet
    dataframe1 = dataframe.active
    
    # Iterate the loop to read the cell values
    for row in range(1, dataframe1.max_row):
        data_row = []
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            data_row.append(col[row].value)
        if len(data_row) == 6:
            id = int(data_row[0])
            name = str(data_row[1])
            x = int(data_row[2])
            y = int(data_row[3])
            w = int(data_row[4])
            h = int(data_row[5])
            val = {}
            val["name"] = name
            val["pose"] = [x, y, w, h]
            map_config[id] = val
    return map_config


def convert_data_raw(path, results, config):
    str = ""
    for key in config.keys():
        if key in results.keys():
            val = results[key]
        else:
            val = {}
            val["val"] = "Nan"
            val["score"]= ""
        str = f'{str}, {val["val"]}, {val["score"]}'
    return str

def write_title(path , config):
    title_first = ","
    title_second = "No., 読み取り時間"
    for key in config.keys():
        title_first = f'{title_first}, {config[key]["name"]}, '
        title_second = f'{title_second}, "AI読み取り結果", "読み取り自信スコア"'
    with open(path, 'a') as file:
        file.write(title_first + "\n")
        file.write(title_second + "\n")

def convert_time(t_detect ):
    
    h = t_detect//3600
    t_detect = t_detect -h*3600
    p = t_detect//60
    s = t_detect%60
    if s<10:
        s=f'0{s}'
    if p<10:
        p=f'0{p}'
    if h<10:
        h=f'0{h}'
    str = f'{h}:{p}:{s}'
    return str
# map_config = read_config_map("./data/map_detector.xlsx")
# results = {1: {'val': '', 'score': 0}, 2: {'val': '146', 'score': 0.8917601108551025}, 3: {'val': '1', 'score': 0.9696356058120728}, 4: {'val': '', 'score': 0}, 5: {'val': '25', 'score': 0.9985036253929138}, 6: {'val': '31', 'score': 0.9988996386528015}, 7: {'val': '9', 'score': 0.8964625597000122}, 8: {'val': '0', 'score': 0.7356021404266357}, 9: {'val': '', 'score': 0}, 10: {'val': '0', 'score': 0.8144099116325378}, 11: {'val': '13', 'score': 0.9951754808425903}, 12: {'val': '140', 'score': 0.9952371716499329}, 13: {'val': '140', 'score': 0.9950234293937683}, 14: {'val': '6', 'score': 0.9555354118347168}, 15: {'val': '131', 'score': 0.9718061685562134}, 16: {'val': '0', 'score': 0.8079658150672913}, 17: {'val': '131', 'score': 0.9694032669067383}, 18: {'val': '0', 'score': 0.9853349328041077}, 19: {'val': '893', 'score': 0.9961639046669006}, 20: {'val': '902', 'score': 0.9291706681251526}, 21: {'val': '2', 'score': 0.9793241620063782}, 22: {'val': '0.00', 'score': 0.9329445362091064}, 23: {'val': '2.58', 'score': 0.9942781329154968}, 24: {'val': '1.19', 'score': 0.995485782623291}, 25: {'val': '0.36', 'score': 0.9943279027938843}}
# # write_ouput_data("test.csv", results, map_config)
# write_title("test.csv",  map_config)
print(convert_time(5400))